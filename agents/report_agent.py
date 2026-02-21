"""
Authors : Pavan R
"""

import os
print(f"[DEBUG] tpc_reports_agent CWD is: {os.getcwd()}")

import re
import math
import json
import pandas as pd
from typing import Callable, Optional, List, Dict, Any
import io
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Any
from rich.pretty import pprint
from langchain_core.messages import HumanMessage
from qgenie.integrations.langchain import QGenieChat
from dvt_insight.connectors.email_agent import send_email
from dvt_insight.core.config import Config
from dvt_insight.tools.db.vector_db_wrapper import VectorDB
from pathlib import Path
from dvt_insight.tools.common.excel_writer import ExcelWriter
from dvt_insight.automation.tools.hws_dvt_tools import HWSDVTTools


class HWSDVTReportAgent:
    
    def get_repo_root(self) -> Path:
        """
        Returns the repository root path by moving up two directories from this file.
        Adjust the number of parents if your repo structure differs.
        """
        return Path(__file__).resolve().parent.parent.parent

    def resolve_relative_path(self, relative_path: str | Path) -> Path:
        """
        Resolves a path relative to the repository root. If already absolute, returns as is.
        """
        p = Path(relative_path)
        if p.is_absolute():
            return p
        return self.get_repo_root() / p
    
    def __init__(self):
        self.vector_db = VectorDB()
        self.metadata_excel_headers = None
        self.meta_data_headers= None
        
        self.tpc_excel_headers = None
        self.evm_excel_headers = None
        self.tpc_headers = None
        self.evm_headers = None
        self.dvt_report_prompt_filepath = self.resolve_relative_path(Config.get("DVT_REPORT_PROMPT_FILE_PATH"))
        self.llm_model = Config.get('LLM_MODEL')
        self.qgenie_api_key = Config.get('QGENIE_API_KEY')
        self.hws_dvt_tools = HWSDVTTools()
        self.report_excel_filepath = self.resolve_relative_path(Config.get("GENERATED_FILE_LOCATION"))
        today = datetime.today().strftime('%Y-%m-%d')
        report_file_name = f"cnss_hws_dvt_report_{today}.xlsx"
        self.excel_file_path = os.path.join(self.report_excel_filepath, report_file_name)
        self.email_id = Config.get('EMAIL_ID')
        
        self.excel_writer = ExcelWriter(self.excel_file_path, Config)
        
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.DEBUG)
    
    def update_markdown_prompt(self, md_filepath, metadata_dict, rows, headers):
        """
        Reads in a Markdown file and returns the content as a single string
        suitable for an LLM prompt, stripping leading/trailing whitespace.
        
        Args:
            md_filepath (str): Path to the markdown (.md) file
        
        Returns:
            str: The cleaned-up LLM prompt string
        """
        try:
            with open(md_filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            # Optionally, normalize line endings
            prompt_template = content.strip().replace('\r\n', '\n')
            # Fill in the template
            prompt = prompt_template.format(
                metadata = metadata_dict,
                records = rows,
                headers = headers,
            )
            return prompt
        except Exception as e:
            print(f"Error reading {md_filepath}: {e}")
            return None
        
    def get_report_headers(self, key, fallback=None):
        """
        Returns the report headers as a list from the env config, using the provided key.
        If not found, returns the fallback list if provided, else raises ValueError.
        """
        headers_str = Config.get(key)
        if headers_str:
            try:
                return json.loads(headers_str)
            except Exception as e:
                raise ValueError(f"Could not parse headers from env var {key}: {e} (was: {headers_str})")
        elif fallback is not None:
            return fallback
        else:
            raise ValueError("No headers found in environment and no fallback provided.")
        

    def build_metadata_table(self, doc):
            """
            Dynamically creates a metadata dictionary from the fields specified in env_config["VECTOR_DB_METADATA_FIELDS"].
            
            meta: Source dictionary containing metadata values.
            env_config: Dictionary holding your environment configuration.
            mget: Function to retrieve a given key from meta safely.
            display_name_map: Optional dict mapping field names to friendly display names.
            
            Returns: dict with keys as field names or mapped display names, and values from meta via mget.
            """
            doc0 = doc.metadata  # Use docs[0]
            meta_keys = [
                "session", "test type", "dvtRef", "buildNum", "SiVersion", "mtp", "nodeType", "tester", "testcaseFile"
            ]
            metadata_dict = {
                "session": doc0.get("test_session_id", "N/A"),
                "test type": doc0.get("test_type", "N/A"),
                "dvtRef": doc0.get("dvtrefnum", "N/A"),
                "buildNum": doc0.get("buildnum", "N/A"),
                "SiVersion": doc0.get("si_version", "N/A"),
                "mtp": doc0.get("mtp", "N/A"),
                "nodeType": doc0.get("node_type", "N/A"),
                "tester": doc0.get("tester", "N/A"),
                "testcaseFile": doc0.get("custome_test_cases_file", "N/A"),
            }
            #metadata_table = "| Parameter | Value |\n|---|---|\n"
            #for key, value in metadata_dict.items():
            #    metadata_table += f"| {key} | {value} |\n"
                
            return metadata_dict
    
    def extract_row(self, doc):
        """
        Given a record (doc), extract required report row as a dict.
        """
        md = doc.metadata
        # You may need to adjust access depending on your doc structure
        # (e.g., doc['channels'][0][...] if channel-specific, or just doc['...'] if flat).

        # Example assumes fields are at top-level or in metadata
        row = {
            "Test Case Name": md.get("testcase_name", "N/A"),
            "Time Stamp": md.get("time_stamp", "N/A"),
            "Test Result Passed": md.get("tpc_Pass", "N/A"),
            "TPC_MAE": md.get("tpc_MAE", "N/A"),
            "RMSE": md.get("tpc_RMSE", "N/A"),
            "TPC Quality Score": md.get("tpc_QualityScore", "N/A"),
            "EVM Score":md.get("evm_Score", "N/A"),
            "EVM Result Passed": md.get("kpi_compliance"),
            "KPI-chEvmDbUL": md.get("chEvmDbUL_kpi", "N/A"),
            "Channel": md.get("channel", "N/A"),
            "DPD": md.get("dpdMode", "N/A"),
            "Temperature": md.get("target_temperature", "N/A"),
            "Phy": md.get("phyId", "N/A"),
            "ChanBW": md.get("cbw", "N/A"),
            "Rate": md.get("rate", "N/A"),
            "RateBW": md.get("rateBw", "N/A"),
            "Nss": md.get("nss", "N/A"),
            "TxMode": md.get("txMode", "N/A"),
            "ChainMask": md.get("chainMask", "N/A"),
            "GuardInterval": md.get("guardInterval", "N/A"),
            "DB Ingestion TS": md.get("ingestion_timestamp", "N/A"),
           
        }
        return row

 

    def extract_sql_from_llm_response(self, llm_output: str) -> str:
        """
        Extracts just the SQL query from LLM output.
        Removes any markdown code block markers (e.g., ```sql) and explanation lines.
        Returns the SQL query as a plain string.
        """
        # Try to extract inside ```sql ... ```
        match = re.search(r"```sql\s*(.*?)```", llm_output, re.DOTALL | re.IGNORECASE)
        if match:
            return match.group(1).strip()
        # Try to extract inside generic ```
        match = re.search(r"```\s*(.*?)```", llm_output, re.DOTALL | re.IGNORECASE)
        if match:
            return match.group(1).strip()
        # As fallback, extract lines starting from SELECT (and until ending semicolon or next empty line/comment)
        lines = llm_output.splitlines()
        sql_lines = []
        in_sql = False
        for line in lines:
            stripped = line.strip()
            if not in_sql and stripped.upper().startswith("SELECT"):
                in_sql = True
            if in_sql:
                if not stripped or stripped.startswith("#"):
                    break
                sql_lines.append(line)
                if ";" in line:
                    break  # Likely end of SQL query
        return "\n".join(sql_lines).strip()

    def evm_conditional_format(self, cell, col_name, cell_value, row_idx):
        col_name = col_name.strip().lower()
        value = str(cell_value).strip().upper()
        if col_name == "test result passed":
            if value == "TRUE":
                cell.fill = PatternFill("solid", fgColor="C6EFCE")  # Green
            elif value == "FALSE":
                cell.fill = PatternFill("solid", fgColor="FFC7CE")  # Red
        elif row_idx % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F3F3F3")

    def extract_table_and_summary(self, response):
        summary_idx = response.find('### Summary:')
        if summary_idx == -1:
            return response.strip(), {}, ""
        
        markdown_table = response[:summary_idx].strip()
        summary_section = response[summary_idx:].strip()

        summary_stats = {}
        analysis_lines = []

        # Parse lines after '### Summary:'
        lines = summary_section.splitlines()[1:]  # skip '### Summary:'
        # Collect simple stats (colon, integer) until first non-matching line
        stats_pattern = re.compile(r'^([A-Za-z0-9 _\-]+):\s*([\d,]+)$')
        stats_done = False

        for line in lines:
            stripped = line.strip()
            if not stats_done:
                match = stats_pattern.match(stripped)
                if match:
                    key, val = match.groups()
                    summary_stats[key.strip()] = val.replace(',', '').strip()
                    continue
                else:
                    stats_done = True  # stats section ended, now in analysis
            # From here on, lines count as analysis (skip empty lines)
            if stripped:
                analysis_lines.append(stripped)

        analysis = "\n".join(analysis_lines)
        return markdown_table, summary_stats, analysis    
    
    def generate_report(self):
        """
        Generates a report based on the latest test execution time.
        """
        vectordb = self.vector_db
        #sql_query_prompt ="""
        #    Generate a PostgreSQL query that selects all columns 
        #    from the table langchain_pg_embedding for the row(s) 
        #    where the cmetadata->>'time_stamp' value is equal to the 
        #    latest (maximum) timestamp among all rows (the most recent entry). 
        #    The time_stamp field is stored inside the JSONB column cmetadata.
        #"""
        #llm_response=self.hws_dvt_tools.llm_call(sql_query_prompt)
        #sql_query=self.extract_sql_from_llm_response(llm_response)
        
        # Run custom query to retrieve all records with the latest test execution time.
        sql_query = """
            SELECT *
            FROM langchain_pg_embedding
            WHERE (cmetadata->>'time_stamp')::timestamp = (
                SELECT MAX((cmetadata->>'time_stamp')::timestamp)
                FROM langchain_pg_embedding
            )
        """

        # 1. Get documents from vectordb
        if vectordb is not None:
            docs = vectordb.run_custom_query(sql_query)
            if not docs:
                self.logger.warning("Vector store query returned no results!")
                return json.dumps({"content": "No records found for the latest test execution time."})
        else:
            self.logger.warning("No vector store available!")
            return json.dumps({"content": "No vector store available!"})

        print(f"[DEBUG] Retrieved {len(docs)} documents from vector store")
        #print("[DEBUG] Sample doc structure:")
        #pprint(docs[0])
        #build meta data : (Top sheet of the report)
        metadata_dict = self.build_metadata_table(docs[0])  # Use docs[0]
        #print(metadata_dict)
        
        # EVM Test report
        headers = [ "Test Case Name", "Time Stamp", "Test Result Passed", "TPC_MAE",
                        "RMSE", "TPC Quality Score", "EVM Score", "EVM Result Passed", "KPI-chEvmDbUL",
                        "Channel", "DPD", "Temperature", "Phy", "ChanBW", "Rate",
                        "RateBW", "Nss", "TxMode", "ChainMask", "GuardInterval",
                        "DB Ingestion TS"]

        # If your extract_evm_row returns a single row, wrap it as a list of lists:
        rows = [self.extract_row(doc) for doc in docs]  # For all docs. 
        
        #Create a prompt for LLM ingestion.
        prompt = self.update_markdown_prompt(self.dvt_report_prompt_filepath, metadata_dict, rows, headers)
        llm_response = self.hws_dvt_tools.llm_call(prompt)
        
        pprint("------Generated Response-------- :")
        print(llm_response)
        pprint("------End Response-------- :")
        
        markdown_table, summary_stats, analysis = self.extract_table_and_summary(llm_response)
        if not analysis:
            analysis = "Analysis not available or not provided. Please check the test output and prompt template."
        metadata_dict.update({
            "Total Test Cases": summary_stats.get("Total Number of Test Cases", "N/A"),
            "Passed": summary_stats.get("Passed", "N/A"),
            "Failed": summary_stats.get("Failed", "N/A"),
        })
        
        #Metadata into excel
        title = "Metadata Summary"              
        report_title = f"HWS DVT Dashboard Test Report"    
        self.excel_writer.add_data_sheet(metadata_dict, title, report_title)
        #Test report summary into Excel
        self.excel_writer.add_table_sheet(
            headers,
            rows,
            sheet_name="DVT Test Report Summary",
            conditional_formats=self.evm_conditional_format,
            autofit=True
        )
        self.excel_writer.save()
                
        return json.dumps({"content": llm_response, "metadata_dict": metadata_dict, "summary_stats": summary_stats, "analysis": analysis})
    
    ##################################################################
    #sendemail with attachment and analysis###########################
    ##################################################################
    def markdown_table_to_html(self, md_table: str) -> str:
        lines = [line.strip() for line in md_table.strip().splitlines() if line.strip()]
        html = '<table>'
        for i, line in enumerate(lines):
            if '---' in line:
                continue
            cells = [cell.strip() for cell in line.split('|') if cell.strip()]
            tag = 'th' if i == 0 else 'td'
            row = ''.join(f'<{tag}>{cell}</{tag}>' for cell in cells)
            html += f'<tr>{row}</tr>'
        html += '</table>'
        return html
    
    def analysis_to_html_bullets_or_paragraph(self, analysis: str) -> str:
        """
        Converts analysis text into HTML.
        - If multiple lines (likely bullet points), wraps each in <li> inside <ul>.
        - If a single paragraph, wraps it in <p>.
        - Returns empty string if input is empty or whitespace.
        """
        lines = [line.strip(" *-") for line in analysis.split('\n') if line.strip()]

        if not lines:
            return ""

        # Heuristic: if more than one line or lines start with bullet-like characters, treat as list
        if len(lines) > 1 or any(analysis.strip().startswith(c) for c in ("-", "*", "•")):
            return "<ul>" + "".join(f"<li>{line}</li>" for line in lines) + "</ul>"
        else:
            return f"<p>{lines[0]}</p>"
        
    from datetime import datetime

    def send_email_with_attachment(self, email_id, attachment, metadata_table, summary_stats, analysis):
        today = datetime.today().strftime('%Y-%m-%d')
        dashboard_image = self.resolve_relative_path("./ui/images/dashboard.png")
        calendar_image = self.resolve_relative_path("./ui/images/calendar.png")
        user_image = self.resolve_relative_path("./ui/images/user.png")
        warning_image = self.resolve_relative_path("./ui/images/warning.png")
        
        print(dashboard_image)
        print(calendar_image)
        print(user_image)
        print(warning_image)
        
        if email_id is None:
            raise ValueError("email_id cannot be None in send_email_with_attachment()")
        if isinstance(email_id, list):
            email_id = ",".join(str(e) for e in email_id)
        elif not isinstance(email_id, str):
            email_id = str(email_id)
        
        # Convert metadata to HTML table
        if isinstance(metadata_table, list):
            metadata_table = "\n".join(metadata_table)
        metadata_html = self.markdown_table_to_html(metadata_table)
        
        # Build analysis HTML
        analysis_html = ""
        if analysis:
            analysis_html = f"""
            <div class="analysis-section">
            <h3 style="margin-top:0;">Analysis</h3>
            {self.analysis_to_html_bullets_or_paragraph(analysis)}
            </div>
            """

        # Build summary stats HTML
        summary_html = ""
        if summary_stats:
            summary_html = '<h3>Summary of Test Results</h3>\n<ul class="summary-list">'
            for k, v in summary_stats.items():
                summary_html += f"<li><strong>{k}:</strong> {v}</li>"
            summary_html += "</ul>"

        body = f"""
        <html>
        <head>
        <meta charset="UTF-8">
        <style>
            body {{
            font-family: 'Segoe UI', Arial, sans-serif;
            background: #f4f6f8;
            color: #344055;
            margin: 0;
            padding: 0;
            }}
            .report-container {{
            background: #fff;
            max-width: 700px;
            margin: 30px auto;
            border-radius: 12px;
            box-shadow: 0 2px 10px rgba(34, 41, 47, 0.08);
            padding: 32px 28px 24px 28px;
            border-top: 8px solid #4B6FA6;
            }}
            .logo-bar {{
            text-align: center;
            padding-bottom: 8px;
            border-bottom: 1px solid #eaeaea;
            }}
            .logo-bar img {{
            max-height: 45px;
            margin-bottom: 4px;
            }}
            h1 {{
            color: #274472;
            font-weight: 700;
            margin-top: 14px;
            font-size: 1.65rem;
            text-align: center;
            }}
            .meta {{
            font-size: 0.96rem;
            color: #788090;
            margin-bottom: 2rem;
            text-align: center;
            }}
            p {{
            margin: 0.8em 0;
            }}
            h3 {{
            color: #486195;
            margin-top: 1.4em;
            margin-bottom: 0.6em;
            font-size: 1.12rem;
            font-weight: 600;
            text-decoration: underline;
            text-underline-offset: 0.18em;
            text-decoration-color: #c2d4f5;
            }}
            ul, ol {{
            margin-left: 1.5em;
            margin-bottom: 1em;
            }}
            table {{
            border-collapse: separate;
            border-spacing: 0;
            width: 100%;
            margin: 18px 0;
            background: #fbfbfd;
            border-radius: 8px;
            overflow: hidden;
            font-size: 0.98rem;
            }}
            th, td {{
            padding: 10px 18px;
            border: none;
            text-align: left;
            }}
            th {{
            background: #e7eef9;
            color: #344055;
            font-weight: 600;
            border-bottom: 2px solid #e0e6ef;
            }}
            tr:nth-child(even) {{
            background: #f5f7fa;
            }}
            tr:hover {{
            background: #e2ebf6;
            transition: background 0.12s;
            }}
            .analysis-section {{
            background: #f7fafc;
            border-left: 4px solid #6f92cd;
            padding: 13px 20px;
            border-radius: 6px;
            margin: 8px 0 18px 0;
            font-size: 0.99rem;
            }}
            .summary-list {{
            padding: 0.8em 1.6em;
            background: #eef3f9;
            border-radius: 8px;
            margin-bottom: 18px;
            }}
            .summary-list li {{
            padding: 3px 0;
            line-height: 1.6;
            }}
            .footer {{
            margin-top: 28px;
            padding-top: 16px;
            border-top: 1px solid #eee;
            font-size: 0.95rem;
            color: #7c828a;
            text-align: center;
            }}
            .auto-note {{
            margin-top: 25px;
            color: #D14343;
            font-size: 0.93rem;
            font-style: italic;
            border-left: 4px solid #ffdfd7;
            background: #fff3f1;
            padding: 10px 18px;
            border-radius: 7px;
            }}
            .icon {{
            vertical-align: middle;
            margin-right: 7px;
            height: 17px;
            width: 17px;
            opacity: 0.65;
            }}
            @media (max-width: 600px) {{
            .report-container {{
                padding: 14px 6px 12px 6px;
            }}
            h1 {{
                font-size: 1.1rem;
            }}
            }}
        </style>
        </head>
        <body>
        <div class="report-container">
            <div class="logo-bar">
            <span>📊 Dashboard</span>
            </div>
            <h1>CNSS HWS DVT Dashboard Test Report</h1>
            <div class="meta">
            <span>🤖 Generated by: QGenie&nbsp;&nbsp;</span>
            <span>📅 Generated on: {today}</span>
            </div>
            <p>Dear Team,</p>
            <p>
            Please find attached the latest <strong>CNSS HWS DVT Dashboard Test Report</strong>.
            </p>
            {summary_html}
            {analysis_html}
            <h3>Report Metadata</h3>
            {metadata_html}
            <h3>Details and Attachments</h3>
            <ul>
            <li>For complete details, including all per-test metrics, pass/fail results and additional information regarding EVM and TPC scores – please <strong>review the attached Excel Report</strong>.</li>
            <li>If you have any questions, please <strong>contact the CNSS HWS AI Dashboard Team</strong>.</li>
            </ul>
            <div class="auto-note">
            <span>
                <span>⚠️</span>
                THIS IS AN AUTOMATED QGENIE REPORT. DO NOT REPLY TO THIS EMAIL.
            </span>
            </div>
            <div class="footer">
            <strong>— CNSS HWS Dashboard Team —</strong>
            </div>
        </div>
        </body>
        </html>
        """

        print(f"DEBUG inside send_email_with_attachment() : email_id = {email_id} | type = {type(email_id)}")
        send_email(email_id, body, attachment)
        
    def dict_to_markdown_table(self, d):
        lines = ['| Key | Value |', '|---|---|']
        for k, v in d.items():
            lines.append(f'| {k} | {v} |')
        return '\n'.join(lines)


def main():
    
    report_agent = HWSDVTReportAgent()
    email_id = report_agent.email_id
    #run the reports agent.  
    result = report_agent.generate_report()
    output = json.loads(result)
    metadata_dict = output["metadata_dict"]
    metadata_table = report_agent.dict_to_markdown_table(metadata_dict)
    summary_stats = output.get("summary_stats", {})
    analysis = output.get("analysis", "")
    excel_file = report_agent.excel_file_path
    
    if email_id is None:
        raise ValueError("email_id cannot be None in send_email_with_attachment()")
    if isinstance(email_id, list):
        email_id = ",".join(str(e) for e in email_id)
    elif not isinstance(email_id, str):
        email_id = str(email_id)

    report_agent.send_email_with_attachment(email_id, str(excel_file), metadata_table, summary_stats, analysis)
        
if __name__ == "__main__":
    main()
    